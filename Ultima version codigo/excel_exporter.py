import os
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def formatear_excel(ruta_archivo, df):
    """Aplica formato profesional al archivo Excel."""
    with pd.ExcelWriter(ruta_archivo, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Productos")
        ws = writer.sheets["Productos"]

        # Estilos
        color_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        fuente_header = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        fuente_datos = Font(name="Calibri", size=11)
        borde = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        align_centro = Alignment(horizontal="center", vertical="center")
        align_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Encabezados
        for col in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col)
            c.font = fuente_header
            c.fill = color_header
            c.alignment = align_centro
            c.border = borde

        # Anchos de columna (agregada "Categoría")
        anchos = {"A": 8, "B": 30, "C": 45, "D": 70, "E": 22}
        for letra, ancho in anchos.items():
            ws.column_dimensions[letra].width = ancho

        # Datos con filas alternadas
        color1 = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        color2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for fila in range(2, len(df) + 2):
            color = color1 if fila % 2 == 0 else color2
            for col in range(1, len(df.columns) + 1):
                c = ws.cell(row=fila, column=col)
                c.font = fuente_datos
                c.border = borde
                c.fill = color
                c.alignment = align_centro if col in (1, 5) else align_izq

        # Congelar encabezados y filtros
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def guardar_excel(productos, resumen_urls, nombre_excel):
    """
    Fase 3: Convertir los datos a DataFrame y guardarlos en un Excel formateado.
    """
    print(f"\n[FASE 3] Guardando todos los productos en UN solo Excel...\n")
    total = len(productos)

    df = pd.DataFrame({
        "N°": range(1, total + 1),
        "Categoría": [p["categoria"] for p in productos],
        "Producto": [p["nombre"] for p in productos],
        "Descripción": [p["descripcion"] for p in productos],
        "Precio": [p["precio"] for p in productos],
    })

    directorio = os.path.dirname(os.path.abspath(__file__))
    ruta_excel = os.path.join(directorio, nombre_excel)
    formatear_excel(ruta_excel, df)

    print(f"  ¡LISTO! Archivo guardado y combinado en:")
    print(f"  {ruta_excel}")
    print(f"\n  Resumen por enlace:")
    print(f"  {'-' * 50}")
    for cat, cant, pags in resumen_urls:
        print(f"    {cat}: {cant} productos ({pags} pág.)")
    print(f"  {'-' * 50}")
    print(f"    TOTAL FINAL: {total} productos")
    print(f"    Columnas: N°, Categoría, Producto, Descripción, Precio")
    print("=" * 60)
