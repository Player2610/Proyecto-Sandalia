import time
import os
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

# ==========================================
# INSTRUCCIONES:
# ==========================================
# Instala las dependencias:
#   python -m pip install selenium webdriver-manager pandas openpyxl
# ==========================================

# ════════════════════════════════════════════════════════
# LISTA DE URLs A SCRAPEAR
# ════════════════════════════════════════════════════════
# Cambia, quita o agrega todos los links que quieras aquí abajo.
URLS = [
    "https://electronilab.co/categoria-producto/wireless/gps/",
    "https://electronilab.co/marca/hobbyking/",
    #"https://electronilab.co/categoria-producto/sensores/",
]

# Nombre del archivo Excel de salida
NOMBRE_EXCEL = "productos_electronilab_completo.xlsx"


def crear_driver():
    """Crea un navegador Chrome controlado por Selenium."""
    opciones = Options()
    opciones.page_load_strategy = 'eager'  # No esperar a que carguen todas las imágenes/scripts
    opciones.add_argument("--headless=new")  # Sin ventana visible
    opciones.add_argument("--disable-gpu")
    opciones.add_argument("--no-sandbox")
    opciones.add_argument("--disable-dev-shm-usage")  # Evitar problemas de memoria con renderer
    opciones.add_argument("--window-size=1920,1080")
    opciones.add_argument("--log-level=3")  # Menos logs
    opciones.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )

    servicio = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=servicio, options=opciones)
    driver.set_page_load_timeout(30)
    return driver


def esperar_productos(driver, timeout=15):
    """Espera a que los productos se carguen en la página (Algolia JS)."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "li.ais-Hits-item, li.product"))
        )
        time.sleep(1)  # Espera extra para que Algolia termine de renderizar
        return True
    except Exception:
        return False


def extraer_productos_pagina(driver):
    """Extrae nombre, precio y link de cada producto en la página actual."""
    productos = []

    # Intentar selectores de Algolia primero, luego WooCommerce estándar
    items = driver.find_elements(By.CSS_SELECTOR, "li.ais-Hits-item")
    if not items:
        items = driver.find_elements(By.CSS_SELECTOR, "li.product")

    for item in items:
        # --- NOMBRE ---
        try:
            titulo = item.find_element(By.CSS_SELECTOR, "h2")
            nombre = titulo.text.strip()
        except Exception:
            nombre = "Sin nombre"

        # --- PRECIO ---
        try:
            # Precio con descuento (dentro de <ins>)
            try:
                precio_el = item.find_element(By.CSS_SELECTOR, "ins .woocommerce-Price-amount")
            except Exception:
                precio_el = item.find_element(By.CSS_SELECTOR, ".woocommerce-Price-amount")
            precio = precio_el.text.strip()
        except Exception:
            try:
                precio_el = item.find_element(By.CSS_SELECTOR, "span.price")
                precio = precio_el.text.strip()
            except Exception:
                precio = "Sin precio"

        # --- LINK ---
        try:
            link_el = item.find_element(By.CSS_SELECTOR, "a[href]")
            link = link_el.get_attribute("href")
        except Exception:
            link = None

        if nombre and nombre != "Sin nombre":
            productos.append({"nombre": nombre, "precio": precio, "link": link})

    return productos


def obtener_descripcion(driver, url):
    """Visita la página individual de un producto y extrae la descripción."""
    try:
        try:
            driver.get(url)
        except TimeoutException:
            pass  # Continuar aunque haya timeout
            
        # Espera breve inicial para que la página renderice estructura básica
        time.sleep(1.5)

        # Descripción corta (intentar primero)
        try:
            desc_corta = driver.find_element(
                By.CSS_SELECTOR, ".woocommerce-product-details__short-description"
            )
            # get_attribute("textContent") obtiene texto, incluso si el elemento no es visible
            texto_corto = desc_corta.get_attribute("textContent").strip()
            if texto_corto and len(texto_corto) > 10:
                return texto_corto
        except Exception:
            pass

        # Descripción larga (pestaña) - Usamos WebDriverWait por si demora en cargar el DOM
        try:
            desc_larga = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#tab-description, .woocommerce-Tabs-panel--description"))
            )
            texto_largo = desc_larga.get_attribute("textContent").strip()
            # Limpiar posible basura de espacios múltiples
            texto_largo = " ".join(texto_largo.split())
            if texto_largo:
                return texto_largo[:600]  # A veces son muy largas, limitamos su tamaño en Excel
        except Exception:
            pass

        # Último intento: meta description del head
        try:
            meta = driver.find_element(By.CSS_SELECTOR, "meta[name='description']")
            texto_meta = meta.get_attribute("content").strip()
            if texto_meta:
                 return texto_meta
        except Exception:
            pass

        return "Sin descripción"
    except Exception:
        return "Sin descripción"


def hay_pagina_siguiente(driver):
    """Verifica si hay un botón 'Siguiente' en la paginación y hace click."""
    try:
        # Selector para Algolia pagination
        next_btn = driver.find_element(
            By.CSS_SELECTOR, "a.ais-Pagination-link[aria-label='Next']"
        )
        if next_btn.is_displayed() and next_btn.is_enabled():
            driver.execute_script("arguments[0].scrollIntoView(true);", next_btn)
            time.sleep(0.5)
            next_btn.click()
            time.sleep(2)  # Esperar que cargue la nueva página
            return True
    except Exception:
        pass

    # Fallback: WooCommerce standard pagination
    try:
        next_link = driver.find_element(By.CSS_SELECTOR, "a.next.page-numbers")
        next_link.click()
        time.sleep(2)
        return True
    except Exception:
        pass

    return False


def formatear_excel(ruta_archivo, df):
    """Aplica formato profesional al archivo Excel."""
    with pd.ExcelWriter(ruta_archivo, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Productos")
        ws = writer.sheets["Productos"]

        # Estilos
        color_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        fuente_header = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        fuente_datos = Font(name="Calibri", size=11)
        borde = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        align_centro = Alignment(horizontal="center", vertical="center")
        align_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Encabezados
        for col in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col)
            c.font = fuente_header
            c.fill = color_header
            c.alignment = align_centro
            c.border = borde

        # Anchos de columna (agregada "Categoría")
        anchos = {"A": 8, "B": 30, "C": 45, "D": 70, "E": 22}
        for letra, ancho in anchos.items():
            ws.column_dimensions[letra].width = ancho

        # Datos con filas alternadas
        color1 = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        color2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for fila in range(2, len(df) + 2):
            color = color1 if fila % 2 == 0 else color2
            for col in range(1, len(df.columns) + 1):
                c = ws.cell(row=fila, column=col)
                c.font = fuente_datos
                c.border = borde
                c.fill = color
                c.alignment = align_centro if col in (1, 5) else align_izq

        # Congelar encabezados y filtros
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def extraer_categoria_de_url(url):
    """Extrae un nombre de categoría legible a partir de la URL."""
    partes = url.rstrip("/").split("/")
    segmentos = [p for p in partes[3:] if p and not p.isdigit() and p != "page"]
    if segmentos:
        return " > ".join(segmentos).replace("-", " ").title()
    return url


def scrape_url_completa(driver, url):
    """Scrapea TODAS las páginas de una URL (paginación incluida)."""
    productos = []
    try:
        driver.get(url)
    except TimeoutException:
        print("  [!] Warning: Timeout al cargar la página. Intentando continuar...")
        
    pagina = 1

    while True:
        print(f"\n    Página {pagina}...", end=" ")

        if not esperar_productos(driver):
            print("(sin productos, fin)")
            break

        prods = extraer_productos_pagina(driver)
        print(f"{len(prods)} productos extraídos")

        if len(prods) == 0:
            break

        for p in prods:
            print(f"      >> {p['nombre'][:55]}  |  {p['precio']}")

        productos.extend(prods)

        if not hay_pagina_siguiente(driver):
            break

        pagina += 1

    return productos, pagina


def main():
    print("=" * 60)
    print("  SCRAPER ELECTRONILAB.CO - MÚLTIPLES URLs")
    print("  Usando Selenium (contenido renderizado por JavaScript)")
    print("=" * 60)
    print(f"\n  URLs configuradas: {len(URLS)}")
    for i, u in enumerate(URLS, 1):
        print(f"    {i}. {u}")

    driver = crear_driver()
    todos = []
    resumen_urls = []

    try:
        # ═══════════════════════════════════════════
        # FASE 1: Recorrer todas las URLs y sus páginas
        # ═══════════════════════════════════════════
        for idx, url in enumerate(URLS, 1):
            categoria = extraer_categoria_de_url(url)
            print(f"\n{'-' * 60}")
            print(f"  [{idx}/{len(URLS)}] Procesando link: {categoria}")
            print(f"  URL: {url}")
            print(f"{'-' * 60}")

            prods, paginas = scrape_url_completa(driver, url)

            # Agregar la categoría a cada producto encontrado
            for p in prods:
                p["categoria"] = categoria

            todos.extend(prods)
            resumen_urls.append((categoria, len(prods), paginas))
            print(f"  Subtotal URL: {len(prods)} productos en {paginas} página(s)")

        total = len(todos)
        print(f"\n{'=' * 60}")
        print(f"  TOTAL ALCANZADO: {total} productos de {len(URLS)} enlace(s)")
        print(f"{'=' * 60}")

        if total == 0:
            print("\n[!] No se encontraron productos en ninguno de los links.")
            return

        # ═══════════════════════════════════════════
        # FASE 2: Obtener descripción de cada producto
        # ═══════════════════════════════════════════
        print(f"\n[FASE 2] Obteniendo descripción de cada producto...\n")

        for i, prod in enumerate(todos, 1):
            if prod["link"]:
                nombre_corto = prod["nombre"][:45]
                print(f"  [{i}/{total}] {nombre_corto}...", end=" ")
                prod["descripcion"] = obtener_descripcion(driver, prod["link"])
                estado = "OK" if prod["descripcion"] != "Sin descripción" else "sin desc."
                print(estado)
                time.sleep(0.5)
            else:
                prod["descripcion"] = "Sin descripción"

        # ═══════════════════════════════════════════
        # FASE 3: Guardar en Excel formateado
        # ═══════════════════════════════════════════
        print(f"\n[FASE 3] Guardando todos los productos en UN solo Excel...\n")

        df = pd.DataFrame({
            "N°": range(1, total + 1),
            "Categoría": [p["categoria"] for p in todos],
            "Producto": [p["nombre"] for p in todos],
            "Descripción": [p["descripcion"] for p in todos],
            "Precio": [p["precio"] for p in todos],
        })

        directorio = os.path.dirname(os.path.abspath(__file__))
        ruta_excel = os.path.join(directorio, NOMBRE_EXCEL)
        formatear_excel(ruta_excel, df)

        print(f"  ¡LISTO! Archivo guardado y combinado en:")
        print(f"  {ruta_excel}")
        print(f"\n  Resumen por enlace:")
        print(f"  {'-' * 50}")
        for cat, cant, pags in resumen_urls:
            print(f"    {cat}: {cant} productos ({pags} pág.)")
        print(f"  {'-' * 50}")
        print(f"    TOTAL FINAL: {total} productos")
        print(f"    Columnas: N°, Categoría, Producto, Descripción, Precio")
        print("=" * 60)

    finally:
        driver.quit()
        print("\n  Navegador cerrado.")


if __name__ == "__main__":
    main()
